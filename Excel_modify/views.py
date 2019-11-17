from django.shortcuts import render
from Excel_modify import forms
import requests
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.conf import settings
from django.http import HttpResponse,Http404
from Excel_modify import models
import openpyxl
result_list=[]
count=True #variable to restrict the resumbmission of form.
    # Create your views here.
    #view to update the file using geocode API
def modify_file(request):
    result_list.clear()
    file=request.FILES['excel_record']
    wb=openpyxl.load_workbook(file)
    sheet=wb.active
    rows=sheet.max_row
    for each in range(2,rows+1):
        try:
            cell_value=str(sheet.cell(row=each,column=1).value)
            adr_value=cell_value.replace(" ","%20")
            app_id="please_assaign your app_id"
            app_code="Please assaign your app"
            adress="https://geocoder.api.here.com/6.2/geocode.json?searchtext="+adr_value+"&app_id="+[kDwDi4jL1kNt6H3YjOK3'+"&app_code=_uPQW7Pwt1oy0JO3aP4gLg&gen=9"
            data=requests.get(adress).json()
            #fetching langitude and Lattitude from API repsonded data
            lat=data["Response"]["View"][0]["Result"][0]["Location"]["NavigationPosition"][0]['Latitude']
            lan=data["Response"]["View"][0]["Result"][0]["Location"]["NavigationPosition"][0]['Longitude']
            result_list.append([cell_value,lat,lan])
        except:
            HttpResponseRedirect("Invalid file/Address details")
    return result_list

#writing the fetched values back to excel
def Excel_writer():
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title="Excel_decoded"
    print("sheet name is renamed as: " + sheet.title)
    sheet.cell(row=1,column=1).value="Adresses"
    sheet.cell(row=1,column=2).value="Lattitude"
    sheet.cell(row=1,column=3).value="Longitude"
    if not len(result_list)==0:
        i=2
        for each in result_list:
            sheet.cell(row=i,column=1).value=each[0]
            sheet.cell(row=i,column=2).value=each[1]
            sheet.cell(row=i,column=3).value=each[2]
            i+=1
        return wb
    return "Excel file is empty"

#view to user to upload the file and list uploaded after submit
def index(request):
    global count
    form=forms.Excel_form()
    message=""
    if request.method=="POST":
        form=forms.Excel_form(request.POST,request.FILES)
        if form.is_valid():
            if count:
                file=modify_file(request)
                count=False   #assaigning False to avoid resumbmission of form.
                if len(file)!=0:
                    message="your file has been processed and listed below"
                    return render(request,'Main_page.html',context={'form':form,'message':message,'files':file}) # sending the API Data to html page to render
                else:
                    message="file doesnt contain any valid adresses/connection problem"
                    return render(request,'Main_page.html',context={'form':form,'message':message})
            count=True#r+eassaigning True  for uploading file after refresh.
            return render(request,'Main_page.html',context={'form':form})
        else:
            return render(request,'Main_page.html',context={'form':form,'message':"Invalid form details"}) #render if form is invalid


    return render(request,'Main_page.html',context={'form':form})

#view to download the file by clicking on the file links
def download(request):
    file_data=Excel_writer()
    response=HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition']='attachment;filename=file.xlsx'
    file_data.save(response)
    return response
